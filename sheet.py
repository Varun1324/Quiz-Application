from openpyxl import Workbook, load_workbook

def insertdatatosheet(name,rollno,score):
    
    file_path = "E:/siva/quizapp/output.xlsx"

    data = [
        ["Name", "Roll No", "Score"],
        [name,rollno,score]
    ]

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(data[0])  # Append headers if creating a new file

    for row in data[1:]:
        sheet.append(row)

    workbook.save(file_path)

    print("Data registered successfully.")
